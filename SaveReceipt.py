import tkinter as tk
import win32com.client
import psutil
import os
import calendar
from openpyxl import workbook, load_workbook
import openpyxl
from openpyxl.descriptors.base import DateTime
from openpyxl.workbook.workbook import Workbook
from datetime import date, datetime

from openpyxl.worksheet import worksheet

wb = load_workbook(
    r'D:\Drive\Documents\Rental Information\Melissa Rental Information\Congress Rent Receipts.xlsx')
ts = wb['Rent Receipt Template']
renterName = ts['B4'].value
rentAmount = ts['B18'].value
rentAmountWords = ts['B20'].value
ws = wb.active
renterInfo = (renterName, rentAmount, rentAmountWords)

currentYear = datetime.now().year
currentMonth = datetime.now().month
currentMonthName = datetime.now().strftime('%B')

fields = ('Renter', 'Rental Amount', 'Rental Amount Words')
root = tk.Tk()
canvas1 = tk.Canvas(root, width=400, height=200)


def makeform(root, fields):
    entries = {}
    rICounter = 0
    for field in fields:
        # print(field)
        row = tk.Frame(root)
        lab = tk.Label(row, width=22, text=field+": ", anchor='w')
        ent = tk.Label(row, width=22, text=renterInfo[rICounter], anchor='w')
        #ent.insert(0, renterInfo[rICounter])
        rICounter += 1
        row.pack(side=tk.TOP,
                 fill=tk.X,
                 padx=5,
                 pady=5)
        lab.pack(side=tk.LEFT)
        ent.pack(side=tk.RIGHT,
                 expand=tk.YES,
                 fill=tk.X)
        entries[field] = ent
    return entries


ents = makeform(root, fields)


def exportToPDF():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = r'D:\Drive\Documents\Rental Information\Melissa Rental Information\Congress Rent Receipts.xlsx'
    wb = o.Workbooks.Open(wb_path)
    ws_index_list = [1]  # say you want to print these sheets
    path_to_pdf = fr'D:\Drive\Documents\Rental Information\Melissa Rental Information\{currentYear}\{currentMonthName} Rent Receipt.pdf'
    wb.WorkSheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    wb.Close()


def getRenterInfo():
    wb = load_workbook(
        r'D:\Drive\Documents\Rental Information\Melissa Rental Information\Congress Rent Receipts.xlsx')
    ts = wb['Rent Receipt Template']
    renterName = ts['B4'].value
    rentAmount = ts['B18'].value
    rentAmountWords = ts['B20'].value
    ws = wb.active
    renterInfo = (renterName, rentAmount, rentAmountWords)


def saveFile():
    source = wb.active
    target = wb.copy_worksheet(source)
    target.title = datetime(currentYear, currentMonth, 1).strftime("%b(%Y)")
    wb.move_sheet(target, -(len(wb.sheetnames)-1))

    target['B12'].value = datetime.now().strftime('%b 1, %Y')
    target['G18'].value = datetime.now().strftime('DL%Y-%m')
    target['B25'].value = f'Rent for {currentMonthName} {currentYear}'
    img = openpyxl.drawing.image.Image('sig.jpg')
    img.anchor = 'H21'
    target.add_image(img)
    wb.active = target
    wb.save(r'D:\Drive\Documents\Rental Information\Melissa Rental Information\Congress Rent Receipts.xlsx')
    label1 = tk.Label(root, text='Saved', fg='green',
                      font=('helvetica', 12, 'bold'))
    canvas1.create_window(200, 75, window=label1)
    exportToPDF()


getRenterInfo()


canvas1.pack()

button1 = tk.Button(text=f'Save {currentMonthName} {currentYear} Receipt',
                    command=saveFile, bg='brown', fg='white')
label2 = tk.Label(root, text=f'Last Receipt {ws.title}', fg='blue', font=(
    'helvetica', 10, 'bold'))
canvas1.create_window(200, 20, window=label2)
canvas1.create_window(200, 50, window=button1)
root.mainloop()

# print(ents)
